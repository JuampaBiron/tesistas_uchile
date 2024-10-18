import logging
import pandas as pd
from openpyxl import load_workbook

# Configuración del logger
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

class GenerateFinalReport:
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
        self.profesores_dictionary = "process_data/profesores.xlsx"
        self.estudiantes = "process_data/estudiantes_postgrados.xlsx"
        self.estudiantes_postgrado = "process_data/estudiantes_postgrado.xlsx"

    def match_files(self):
        logging.info("Iniciando la carga de archivos Excel.")
        try:
            df_profesores = pd.read_excel(self.profesores_dictionary)[['Nombre portafolio', "Nombre ucampus", 'Sexo', 'Rut', 'Jornada', 'Jerarquia']]
            df_estudiantes = pd.read_excel(self.estudiantes)[["Nombre estudiante", "Rut", "Cohorte", "Programa", "Tesista", "Profesor guia"]]
            logging.info("Archivos cargados exitosamente.")
        except Exception as e:
            logging.error("Error al cargar los archivos: %s", e)
            return

        logging.info("Realizando la combinación de DataFrames.")
        df_raw_resultado = df_estudiantes.merge(
            df_profesores,
            left_on='Profesor guia',
            right_on='Nombre ucampus',
            how='left'
        )

        # Eliminar las columnas auxiliares si es necesario
        df_raw_resultado = df_raw_resultado.drop(columns=['Rut_x', 'Profesor guia', 'Nombre ucampus'])
        df_raw_resultado = df_raw_resultado.rename(columns={
            "Rut_y": "Rut Prof. Guía",
            "Jerarquia": "Jerarquía",
            "Nombre portafolio": "Nombre Prof. Guía"
        })
        
        logging.info("Combinación de DataFrames completa.")

        df_tesistas = df_raw_resultado.loc[df_raw_resultado['Tesista'] == True]
        df_regulares = df_raw_resultado.loc[df_raw_resultado['Tesista'] == False]

        # Crear un DataFrame para la hoja de README
        readme_data = {
            "Descripción": [
                "Este archivo contiene información sobre estudiantes de postgrado.",
                "La hoja 'tesistas' incluye estudiantes que están realizando su tesis. Los que tienen profesor guía son los que lo tienen declarado directamente en su BIA, los que no los tienen, son casos en los que existe inscripción de tesis asociada a su plan declarada en la BIA.",
                "La hoja 'regulares' incluye estudiantes que no cumplen con las reglas anteriores."
            ]
        }
        df_readme = pd.DataFrame(readme_data)

        logging.info("Guardando resultados en el archivo Excel.")
        try:
            with pd.ExcelWriter(self.estudiantes_postgrado) as writer:
                df_readme.to_excel(writer, index=False, sheet_name="README")
                df_tesistas.to_excel(writer, index=False, sheet_name="tesistas")
                df_regulares.to_excel(writer, index=False, sheet_name="regulares")
            logging.info("Resultados guardados exitosamente.")
        except Exception as e:
            logging.error("Error al guardar el archivo de Excel: %s", e)
            return

        # Cargar el libro de trabajo para autoajustar columnas
        wb = load_workbook(self.estudiantes_postgrado)

        # Función para autoajustar las columnas
        def auto_adjust_columns(sheet):
            for column in sheet.columns:
                max_length = 0
                column = [cell for cell in column]  # Convertir la columna en una lista
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Ajuste de 2 para un poco de espacio
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Autoajustar las columnas y agregar filtros en cada hoja
        logging.info("Autoajustando columnas y agregando filtros.")
        def add_filters(sheet):
            sheet.auto_filter.ref = sheet.dimensions  # Aplicar filtros a todas las columnas

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            auto_adjust_columns(ws)
            add_filters(ws)

        # Guardar los cambios en el archivo
        wb.save(self.estudiantes_postgrado)
        logging.info("Cambios guardados en el archivo Excel.")

    def run_workflow(self):
        logging.info(f"************ Inicio del workflow {self.__class__.__name__} ************")
        try:
            self.match_files()
        except Exception as e:
            logging.error(f"Error en {self.__class__.__name__}: %s", e)
        finally:
            logging.info(f"************ Termino del workflow {self.__class__.__name__} ************")

if __name__ == "__main__":
    proceso = GenerateFinalReport()
    proceso.run_workflow()
